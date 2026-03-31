"""
SBR Tracker Automation — Production Streamlit App
Phases 1–4 · Steps 1–34
Professional dark dashboard UI
"""

import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date, timedelta
import io
import warnings
import math
import traceback

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG  — must be first Streamlit call
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="SBR Tracker Automation",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Max upload size set here in code so config.toml is NOT required ──────────
# This means uploading all 3 files to GitHub without any sub-folders works fine.
# Streamlit reads this before the uploader widgets are rendered.
st.config.set_option("server.maxUploadSize", 500)

MAX_FILE_BYTES = 500 * 1024 * 1024  # 500 MB — enforced in Python as well

# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', sans-serif !important; }
.stApp { background: #0d0f14 !important; }

section[data-testid="stSidebar"] {
    background: #0b0d12 !important;
    border-right: 1px solid #1c2030 !important;
}
section[data-testid="stSidebar"] > div { padding-top: 0 !important; }

#MainMenu, footer, header { visibility: hidden; }
.block-container { padding: 0 !important; max-width: 100% !important; }

/* NAVBAR */
.sbr-navbar {
    display:flex; align-items:center; justify-content:space-between;
    padding:0 32px; height:56px; background:#0d0f14;
    border-bottom:1px solid #1c2030;
}
.sbr-brand { display:flex; align-items:center; gap:10px; }
.sbr-brand-icon {
    width:30px; height:30px;
    background:linear-gradient(135deg,#3b82f6,#6366f1);
    border-radius:8px; display:flex; align-items:center;
    justify-content:center; font-size:11px; font-weight:700; color:#fff;
}
.sbr-brand-name { font-size:15px; font-weight:600; color:#fff; letter-spacing:-0.3px; }
.sbr-brand-chip {
    font-size:10px; color:#4a5270; background:#13161f;
    border:1px solid #1c2030; padding:2px 8px; border-radius:4px;
    font-family:'DM Mono',monospace;
}
.sbr-nav-right { display:flex; align-items:center; gap:16px; }
.sbr-pill { display:flex; align-items:center; gap:5px; font-size:12px; color:#4a5270; }
.sbr-dot { width:6px; height:6px; border-radius:50%; }
.sbr-dot.green { background:#22c55e; }
.sbr-dot.amber { background:#f59e0b; }

/* SIDEBAR */
.sb-lbl {
    font-size:10px; font-weight:600; letter-spacing:1.1px;
    text-transform:uppercase; color:#2e3448; margin:16px 0 7px 2px;
}
.phase-row {
    display:flex; align-items:center; gap:9px;
    padding:8px 8px; border-radius:8px; margin-bottom:2px;
}
.phase-row.done   { background:#0b1a10; }
.phase-row.active { background:#101828; }
.pnum {
    width:22px; height:22px; border-radius:6px;
    display:flex; align-items:center; justify-content:center;
    font-size:10px; font-weight:700; font-family:'DM Mono',monospace; flex-shrink:0;
}
.pnum.done   { background:#0f2318; color:#4ade80; }
.pnum.active { background:#0f1e36; color:#60a5fa; }
.pnum.idle   { background:#13161f; color:#2e3448; }
.pnum.error  { background:#2a0f0f; color:#f87171; }
.pinfo { flex:1; min-width:0; }
.ptitle { font-size:12px; font-weight:500; color:#8892a4; line-height:1.3; }
.phase-row.done .ptitle, .phase-row.active .ptitle { color:#e2e4e9; }
.psub { font-size:10px; color:#2e3448; margin-top:1px; }
.pbadge {
    font-size:9px; font-family:'DM Mono',monospace;
    padding:2px 6px; border-radius:4px; flex-shrink:0;
}
.pbadge.done   { background:#0f2318; color:#4ade80; }
.pbadge.active { background:#0f1e36; color:#60a5fa; }
.pbadge.idle   { background:#13161f; color:#2e3448; }
.pbadge.error  { background:#2a0f0f; color:#f87171; }

.fs-row { display:flex; align-items:center; gap:7px; padding:5px 6px; border-radius:5px; margin-bottom:1px; }
.fs-dot { width:6px; height:6px; border-radius:50%; flex-shrink:0; }
.fs-dot.ok   { background:#22c55e; }
.fs-dot.miss { background:#1c2030; border:1px solid #2e3448; }
.fs-name { font-size:11px; color:#4a5270; flex:1; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.fs-name.ok { color:#8892a4; }
.fs-sz { font-size:10px; color:#2e3448; font-family:'DM Mono',monospace; flex-shrink:0; }

/* MAIN */
.main-hdr { padding:22px 32px 18px; border-bottom:1px solid #1c2030; }
.main-title { font-size:20px; font-weight:600; color:#fff; letter-spacing:-0.5px; margin-bottom:3px; }
.main-sub   { font-size:12px; color:#4a5270; }

/* PIPELINE */
.pip-wrap { padding:18px 32px 0; }
.pip-track { display:flex; align-items:center; }
.pip-dot { width:10px; height:10px; border-radius:50%; flex-shrink:0; }
.pip-dot.done   { background:#4ade80; }
.pip-dot.active { background:#3b82f6; box-shadow:0 0 0 4px rgba(59,130,246,.18); }
.pip-dot.idle   { background:#1c2030; border:1px solid #2e3448; }
.pip-line { flex:1; height:1px; }
.pip-line.done { background:#4ade80; opacity:.3; }
.pip-line.idle { background:#1c2030; }
.pip-labels { display:flex; justify-content:space-between; padding:5px 0 0; }
.pip-lbl { font-size:10px; color:#2e3448; text-align:center; min-width:44px; }
.pip-lbl.done   { color:#4ade80; }
.pip-lbl.active { color:#60a5fa; }

/* METRICS */
.metrics-row { display:grid; grid-template-columns:repeat(4,1fr); gap:10px; padding:16px 32px; }
.mc { background:#0f1118; border:1px solid #1c2030; border-radius:10px; padding:14px 16px; }
.mc-lbl { font-size:10px; color:#3d4460; text-transform:uppercase; letter-spacing:.8px; margin-bottom:5px; }
.mc-val { font-size:22px; font-weight:600; color:#fff; font-family:'DM Mono',monospace; letter-spacing:-1px; }
.mc-val.blue  { color:#60a5fa; }
.mc-val.green { color:#4ade80; }
.mc-val.amber { color:#fbbf24; }
.mc-delta { font-size:10px; margin-top:4px; color:#4a5270; }
.mc-delta.info { color:#60a5fa; }
.mc-delta.up   { color:#4ade80; }

/* UPLOAD ZONE */
.uz-outer {
    margin:4px 32px 14px; background:#0b0d12;
    border:1px solid #1c2030; border-radius:12px; padding:16px 18px;
}
.uz-title { font-size:10px; font-weight:600; letter-spacing:1px; text-transform:uppercase; color:#2e3448; margin-bottom:12px; }
.uz-grid { display:grid; grid-template-columns:1fr 1fr; gap:8px; }
.uz-card {
    display:flex; align-items:center; gap:10px;
    background:#0f1118; border:1.5px dashed #1c2030;
    border-radius:9px; padding:11px 13px;
}
.uz-card.ok { border:1px solid #132a1a; background:#0b1a0f; }
.uz-card.wide { grid-column:span 2; }
.uz-type {
    width:30px; height:30px; border-radius:7px;
    display:flex; align-items:center; justify-content:center;
    font-size:10px; font-weight:700; flex-shrink:0;
}
.uz-type.xl  { background:#0f2318; color:#4ade80; }
.uz-type.csv { background:#0f1e36; color:#60a5fa; }
.uz-type.emp { background:#13161f; color:#2e3448; font-size:18px; font-weight:300; }
.uz-info { flex:1; min-width:0; }
.uz-name { font-size:12px; font-weight:500; color:#3d4460; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.uz-card.ok .uz-name { color:#c0c6d2; }
.uz-meta { font-size:10px; color:#2e3448; margin-top:2px; font-family:'DM Mono',monospace; }
.uz-card.ok .uz-meta { color:#4ade80; }
.uz-chk {
    width:16px; height:16px; border-radius:50%;
    background:#0f2318; color:#4ade80;
    display:flex; align-items:center; justify-content:center; font-size:9px; flex-shrink:0;
}

/* RUN BUTTON */
.run-outer { padding:0 32px 14px; }
.run-outer .stButton > button {
    width:100% !important;
    background:linear-gradient(135deg,#2563eb,#4f46e5) !important;
    color:#fff !important; border:none !important;
    border-radius:10px !important; padding:14px 24px !important;
    font-size:14px !important; font-weight:600 !important;
    font-family:'DM Sans',sans-serif !important;
    letter-spacing:-0.2px !important; height:auto !important;
}
.run-outer .stButton > button:hover { opacity:.9 !important; border:none !important; }
.run-outer .stButton > button:disabled {
    background:#1c2030 !important; color:#3d4460 !important;
}

/* LOG */
.log-panel {
    margin:0 32px 14px; background:#080a0f;
    border:1px solid #1c2030; border-radius:10px; overflow:hidden;
}
.log-hdr {
    display:flex; align-items:center; gap:8px;
    padding:9px 14px; border-bottom:1px solid #1c2030; background:#0d0f14;
}
.log-hdr-dot { width:7px; height:7px; border-radius:50%; background:#4ade80; flex-shrink:0; }
.log-hdr-title { font-size:11px; color:#4a5270; font-weight:500; flex:1; }
.pulse { width:6px; height:6px; border-radius:50%; background:#3b82f6;
         animation:pls 1.3s ease-in-out infinite; flex-shrink:0; }
@keyframes pls { 0%,100%{opacity:1;transform:scale(1);}50%{opacity:.3;transform:scale(.7);} }
.prun-txt { font-size:11px; color:#3b82f6; }
.log-body {
    padding:12px 14px; font-family:'DM Mono',monospace;
    font-size:11px; line-height:1.9; max-height:260px; overflow-y:auto;
}
.ll { display:flex; gap:14px; }
.lt { color:#1e2538; flex-shrink:0; }
.lm { color:#5a6280; }
.lm.ok    { color:#4ade80; }
.lm.info  { color:#60a5fa; }
.lm.ph    { color:#a78bfa; font-weight:500; }
.lm.warn  { color:#fbbf24; }
.lm.error { color:#f87171; }

/* DOWNLOAD */
.dl-outer { padding:0 32px 32px; }
.dl-title { font-size:10px; font-weight:600; letter-spacing:1px; text-transform:uppercase; color:#2e3448; margin-bottom:10px; }
.dl-grid { display:grid; grid-template-columns:repeat(5,1fr); gap:8px; }
.dl-card {
    background:#0f1118; border:1px solid #1c2030;
    border-radius:10px; padding:13px; display:flex;
    flex-direction:column; gap:5px; position:relative; overflow:hidden;
}
.dl-card::after { content:''; position:absolute; bottom:0; left:0; right:0; height:2px; }
.dl-card.ready        { border-color:#132a1a; }
.dl-card.ready::after { background:linear-gradient(90deg,#4ade80,#22c55e); }
.dl-card.final        { border-color:#13203a; }
.dl-card.final::after { background:linear-gradient(90deg,#3b82f6,#6366f1); }
.dl-card.error        { border-color:#2a0f0f; }
.dl-card.error::after { background:linear-gradient(90deg,#ef4444,#dc2626); }
.dl-lbl { font-size:9px; color:#2e3448; text-transform:uppercase; letter-spacing:.8px; }
.dl-card.ready .dl-lbl, .dl-card.final .dl-lbl { color:#4a5270; }
.dl-ico { font-size:20px; margin:2px 0; }
.dl-name { font-size:11px; font-weight:500; color:#3d4460; }
.dl-card.ready .dl-name { color:#b0b6c4; }
.dl-card.final .dl-name { color:#93c5fd; }
.dl-card.error .dl-name { color:#f87171; }
.dl-sz { font-size:10px; color:#2e3448; font-family:'DM Mono',monospace; }
.dl-card.ready .dl-sz { color:#3d4460; }
.dl-wait { font-size:10px; color:#1e2538; font-style:italic; margin-top:2px; }
.dl-outer .stDownloadButton > button {
    width:100% !important; background:#0f2318 !important;
    color:#4ade80 !important; border:1px solid #1a4028 !important;
    border-radius:6px !important; font-size:11px !important;
    font-weight:600 !important; padding:5px 8px !important;
    height:auto !important; margin-top:4px !important;
}

/* SCROLLBAR */
::-webkit-scrollbar { width:4px; height:4px; }
::-webkit-scrollbar-track { background:#0d0f14; }
::-webkit-scrollbar-thumb { background:#1c2030; border-radius:4px; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
_keys = ["p1_out","p2_out","p3_out","p4_out",
         "p1_stats","p2_stats","p3_stats","p4_stats",
         "run_log","error_phase"]
for k in _keys:
    if k not in st.session_state:
        st.session_state[k] = [] if k == "run_log" else None


# ─────────────────────────────────────────────────────────────────────────────
# PURE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
FMT_CURRENCY = '$#,##0.00'
FMT_DATE     = 'MM/DD/YYYY'
FMT_TEXT     = '@'
FMT_GENERAL  = 'General'

COLUMN_FORMATS = {
    1:FMT_GENERAL,2:FMT_GENERAL,3:FMT_TEXT,4:FMT_GENERAL,
    5:FMT_GENERAL,6:FMT_TEXT,7:FMT_TEXT,8:FMT_DATE,
    9:FMT_DATE,10:FMT_TEXT,11:FMT_CURRENCY,12:FMT_CURRENCY,
    13:FMT_CURRENCY,14:FMT_CURRENCY,15:FMT_CURRENCY,16:FMT_TEXT,
    17:FMT_TEXT,18:FMT_TEXT,19:FMT_TEXT,20:FMT_TEXT,
    21:FMT_TEXT,22:FMT_DATE,23:FMT_GENERAL,24:FMT_TEXT,
    25:FMT_TEXT,26:FMT_TEXT,27:FMT_TEXT,28:FMT_TEXT,
    29:FMT_DATE,30:FMT_TEXT,31:FMT_DATE,32:FMT_TEXT,
    33:FMT_TEXT,34:FMT_DATE,35:'mm/dd/yyyy',36:FMT_TEXT,37:FMT_TEXT,
}

MIN_VALID_DATE = pd.Timestamp('2000-01-01')

PROCESS_STATUSES = {
    'bill resubmitted','sbr time lapsed','payment eor cases','ppo reduction',
    'sbr sent after time lapsed','billing submission timeline expired','settled with agreement',
}
PROTECTED_STATUSES = {'sbr sent','study closed','sbr in queue','pending sbr'}
VALID_SOL_STATUSES = {'pending sbr','sbr sent','study closed'}

def copy_font(f):
    return Font(name=f.name or 'Calibri', size=f.size or 11, bold=f.bold, italic=f.italic,
                color=f.color.rgb if f.color and f.color.type=='rgb' else '000000')
def copy_fill(f):
    if f and f.fill_type and f.fill_type != 'none':
        try: return PatternFill(fill_type=f.fill_type,
                start_color=f.start_color.rgb if f.start_color else 'FFFFFF',
                end_color=f.end_color.rgb if f.end_color else 'FFFFFF')
        except: pass
    return PatternFill(fill_type=None)
def copy_border(b):
    def s(x): return Side(style=x.style, color=x.color.rgb if x.color and x.color.type=='rgb' else '000000') if x and x.style else Side()
    return Border(left=s(b.left), right=s(b.right), top=s(b.top), bottom=s(b.bottom))
def copy_alignment(a):
    return Alignment(horizontal=a.horizontal or 'center', vertical=a.vertical or 'center', wrap_text=a.wrap_text)

def get_row_format_from_existing(ws, sample_row=2):
    ef = {}
    for col in range(1, 38):
        c = ws.cell(row=sample_row, column=col)
        ef[col] = {'font':copy_font(c.font),'fill':copy_fill(c.fill),
                   'border':copy_border(c.border),'alignment':copy_alignment(c.alignment),
                   'number_format':c.number_format}
    return ef

def apply_formatting_to_row(ws, row_idx, ef):
    for col in range(1, 38):
        cell = ws.cell(row=row_idx, column=col)
        f = ef.get(col, {})
        if f.get('font'):   cell.font   = f['font']
        if f.get('fill'):   cell.fill   = f['fill']
        if f.get('border'): cell.border = f['border']
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=f.get('alignment', Alignment()).wrap_text)
        df = COLUMN_FORMATS.get(col)
        if df:                  cell.number_format = df
        elif f.get('number_format'): cell.number_format = f['number_format']

def normalize(val):
    return '' if val is None else str(val).strip().lower()
def is_paper_type(val):
    return normalize(val) in ('paper','paper attorney billed','paper employer billed')
def is_electronic_or_blank(val):
    return normalize(val) in ('','electronic')
def safe_date(val):
    if val is None: return None
    try:
        if pd.isna(val): return None
    except: pass
    try: return pd.to_datetime(val).date()
    except: return None
def to_comparable_date(val):
    if val is None: return None
    s = str(val).strip()
    if s in ('','nan','NaT','None','NA','N/A'): return None
    try: return pd.to_datetime(s).date()
    except: return None
def safe_to_datetime(val):
    if val is None: return None
    if isinstance(val,(int,float)) and val==0: return None
    s = str(val).strip()
    if s in ('','0'): return None
    try:
        ts = pd.to_datetime(val)
        return None if ts < MIN_VALID_DATE else ts
    except: return None
def to_python_datetime(ts):
    return datetime(ts.year, ts.month, ts.day)
def to_date(val):
    if val is None: return None
    if isinstance(val, datetime): return val.date()
    if isinstance(val, date): return val
    try: return pd.to_datetime(val, format='mixed', dayfirst=False).date()
    except: return None
def to_date_obj(val):
    if val is None: return None
    if isinstance(val, datetime): return val.replace(hour=0,minute=0,second=0,microsecond=0)
    if isinstance(val, date): return datetime(val.year,val.month,val.day,0,0,0)
    try:
        dt = pd.to_datetime(val).to_pydatetime()
        return dt.replace(hour=0,minute=0,second=0,microsecond=0)
    except: return None
def to_number(val):
    if val is None: return None
    try: return float(val)
    except: return None
def write_date_cell(cell, d):
    cell.value = datetime(d.year,d.month,d.day,0,0,0)
    cell.number_format = 'M/D/YYYY'
    cell.alignment = Alignment(horizontal='center')
def contains_eob(z):
    return any(kw in normalize(z) for kw in ['eob zero','eob payment'])
def get_col_idx(headers, name):
    try: return headers.index(name)+1
    except: raise ValueError(f"Column '{name}' not found in headers.")
def first_of_current_month():
    t = date.today(); return date(t.year,t.month,1)
def fmt_bytes(n):
    if n is None: return "—"
    if n<1024: return f"{n} B"
    if n<1024**2: return f"{n/1024:.1f} KB"
    if n<1024**3: return f"{n/1024**2:.1f} MB"
    return f"{n/1024**3:.1f} GB"
def ts():
    return datetime.now().strftime("%H:%M:%S")
def log_line(log, msg, cls=""):
    log.append(f'<div class="ll"><span class="lt">{ts()}</span><span class="lm {cls}">{msg}</span></div>')


# ─────────────────────────────────────────────────────────────────────────────
# PHASE RUNNERS
# ─────────────────────────────────────────────────────────────────────────────

def run_phase1(tracker_bytes, bdr_bytes, log):
    log_line(log, "━━ PHASE 1 · STEPS 1–6 ━━━━━━━━━━━━━━━━━━", "ph")
    bdr     = pd.read_csv(io.BytesIO(bdr_bytes))
    tracker = pd.read_excel(io.BytesIO(tracker_bytes), sheet_name="Timeline")
    log_line(log, f"✓ BDR loaded · {len(bdr):,} rows", "ok")
    log_line(log, f"✓ Master-Tracker loaded · {len(tracker):,} rows", "ok")

    existing_ids  = set(tracker['Study Id'].dropna().astype(str))
    bdr['_sid']   = bdr['STUDY_ID'].astype(str)
    new_cases     = bdr[~bdr['_sid'].isin(existing_ids)].copy()
    new_cases.drop('_sid', axis=1, inplace=True)
    log_line(log, f"→ New cases identified: {len(new_cases):,}", "info")

    wb = load_workbook(io.BytesIO(tracker_bytes))
    ws = wb["Timeline"]
    next_row = len(tracker) + 2
    ef = get_row_format_from_existing(ws, sample_row=2)

    if len(new_cases) > 0:
        for _, case in new_cases.iterrows():
            ws.cell(row=next_row,column=1).value = case['STUDY_ID']
            ws.cell(row=next_row,column=2).value = case['CASEID']
            ws.cell(row=next_row,column=3).value = case['PATIENT_NAME']
            ws.cell(row=next_row,column=4).value = case['PATIENTREGID']
            ws.cell(row=next_row,column=5).value = case.get('EXTERNAL MRN','')
            ws.cell(row=next_row,column=6).value = case['BUSINESSNAME']
            ws.cell(row=next_row,column=7).value = case['INSURANCE NAME']
            ws.cell(row=next_row,column=8).value = pd.to_datetime(case['DATEOFSERVICE']).date() if pd.notna(case['DATEOFSERVICE']) else None
            ws.cell(row=next_row,column=9).value = pd.to_datetime(case['BILLDATE']).date() if pd.notna(case['BILLDATE']) else None
            next_row += 1
        log_line(log, f"✓ {len(new_cases):,} new cases written (cols A–I)", "ok")

    bdr_lookup = bdr.set_index('STUDY_ID').to_dict('index')
    updated = 0
    for row_idx in range(2, next_row):
        sid = ws.cell(row=row_idx,column=1).value
        if sid in bdr_lookup:
            d = bdr_lookup[sid]
            for col,key in zip(range(10,20),['APPOINTMENTTYPE','AMOUNTBILLED','PAID','WRITEOFF',
                                              'OUTSTANDING','OMFS','OWNER_NAME','PORTFOLIO_NAME',
                                              'SUB OWNER_NAME','STATUS']):
                ws.cell(row=row_idx,column=col).value = d.get(key,'')
            updated += 1
    log_line(log, f"✓ {updated:,} rows updated (cols J–S)", "ok")

    if len(new_cases) > 0:
        ns = len(tracker) + 2; ne = ns + len(new_cases) - 1
        for r in range(ns, ne+1):
            ws.cell(row=r,column=20).value = "Yes"
            ws.cell(row=r,column=24).value = f'=IF(W{r}>=30,"30 days passed","Under 30 Days")'
            ws.cell(row=r,column=25).value = f'=IF(OR(AB{r}="Yes",AD{r}="Yes"),"Response Received","No Response")'
            ws.cell(row=r,column=27).value = (f'=IF(X{r}="Under 30 Days","No Action Required",'
                                               f'IF(T{r}="No","No Action Required",'
                                               f'IF(Y{r}="No Response","Send No Response Letter","Response Received")))')
            ws.cell(row=r,column=28).value = "No"
            ws.cell(row=r,column=30).value = "No"
            ws.cell(row=r,column=32).value = "No"
            ws.cell(row=r,column=33).value = "Letter not Sent"
            ws.cell(row=r,column=35).value = f'=IF(Y{r}="Response Received",MAX(AE{r},AC{r})+60,DATE(1900,1,1))'
            ws.cell(row=r,column=36).value = f'=IF(Y{r}="Response Received","Yes","No")'
            ws.cell(row=r,column=37).value = "Under Billing Cycle"
            apply_formatting_to_row(ws, r, ef)
        log_line(log, f"✓ Excel formulas applied to {len(new_cases):,} new rows", "ok")

    closed = 0
    for row_idx in range(2, next_row):
        if ws.cell(row=row_idx,column=19).value == "CLOSE":
            ws.cell(row=row_idx,column=37).value = "Study Closed"; closed += 1
    log_line(log, f"✓ {closed:,} cases marked Study Closed", "ok")
    log_line(log, "✓ PHASE 1 COMPLETE → output ready for download", "ok")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read(), {"new_cases":len(new_cases),"updated":updated,"closed":closed,"total_rows":next_row-2}


def run_phase2(tracker_bytes, submission_bytes, log):
    log_line(log, "━━ PHASE 2 · STEPS 7–15 ━━━━━━━━━━━━━━━━━", "ph")
    tracker_df    = pd.read_excel(io.BytesIO(tracker_bytes), sheet_name="Timeline")
    total_rows    = tracker_df.shape[0]
    last_data_row = total_rows + 1

    try:
        submission = pd.read_csv(io.BytesIO(submission_bytes))
        log_line(log, f"✓ Submission loaded · CSV · {len(submission):,} rows", "ok")
    except Exception:
        xl = pd.ExcelFile(io.BytesIO(submission_bytes))
        exp = ['2021 to 2024','2025-2026']
        found = [s for s in exp if s in xl.sheet_names]
        if found:
            submission = pd.concat([xl.parse(s) for s in found], ignore_index=True)
            log_line(log, f"✓ Submission loaded · {len(found)} sheets · {len(submission):,} rows", "ok")
        else:
            submission = xl.parse(xl.sheet_names[0])
            log_line(log, f"✓ Submission loaded · {len(submission):,} rows", "ok")

    lkp = (submission.drop_duplicates(subset='STUDY_ID',keep='last')
           .assign(STUDY_ID=lambda df: df['STUDY_ID'].astype(str).str.strip())
           .set_index('STUDY_ID')[['EDI Service Type','Submission Date']]
           .to_dict('index'))
    log_line(log, f"✓ Lookup built · {len(lkp):,} unique Study IDs", "ok")

    wb = load_workbook(io.BytesIO(tracker_bytes)); ws = wb["Timeline"]
    ws.insert_cols(23, 3)

    matched = not_found = 0
    for r in range(2, last_data_row+1):
        sid = ws.cell(row=r,column=1).value
        if sid is None: continue
        k = str(sid).strip()
        if k in lkp:
            sub = lkp[k]; raw = sub.get('EDI Service Type'); dv = safe_date(sub.get('Submission Date'))
            clean = None
            if raw is not None:
                try:
                    if not math.isnan(float(raw)): clean = str(raw).strip() or None
                except (ValueError,TypeError): clean = str(raw).strip() or None
            if clean is None and dv is None: not_found += 1
            else: ws.cell(row=r,column=23).value=clean; ws.cell(row=r,column=24).value=dv; matched += 1
        else: not_found += 1

    log_line(log, f"✓ Submission data pulled · {matched:,} rows matched", "ok")

    for r in range(2, last_data_row+1):
        if ws.cell(row=r,column=24).value is not None:
            ws.cell(row=r,column=25).value = f'=X{r}=V{r}'
    log_line(log, "→ Comparison formula written to temp Y column", "info")

    uv = am = na = pv = 0
    for r in range(2, last_data_row+1):
        tw=ws.cell(row=r,column=23).value; tx=ws.cell(row=r,column=24).value
        ev=ws.cell(row=r,column=22).value; es=str(ev).strip() if ev is not None else ''
        is_na=es in ('nan','NaT','None','NA','N/A'); is_bl=ev is None or es==''
        has_d=not is_bl and not is_na
        if tw is None and tx is None:
            if has_d: pv+=1
            continue
        nd=to_comparable_date(tx)
        if is_na: na+=1
        elif is_bl or nd!=to_comparable_date(ev): ws.cell(row=r,column=21).value=tw; ws.cell(row=r,column=22).value=tx; uv+=1
        else: am+=1

    ws.delete_cols(23, 3)
    log_line(log, f"✓ Temp columns removed · U&V updated: {uv:,}", "ok")

    lag=elec=paper=0
    for r in range(2, last_data_row+1):
        bst=ws.cell(row=r,column=21).value
        if normalize(bst)!='': ws.cell(row=r,column=23).value=f'=TODAY()-V{r}'; lag+=1
        if   is_electronic_or_blank(bst): ws.cell(row=r,column=24).value=f'=IF(W{r}>=30,"30 days passed","Under 30 Days")'; elec+=1
        elif is_paper_type(bst):          ws.cell(row=r,column=24).value=f'=IF(W{r}>=45,"45 days passed","Under 45 Days")'; paper+=1

    log_line(log, f"✓ Lag Time formula applied · {lag:,} rows", "ok")
    log_line(log, f"✓ 30-day: {elec:,} · 45-day (paper): {paper:,}", "ok")
    log_line(log, "✓ PHASE 2 COMPLETE → output ready for download", "ok")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read(), {"matched":matched,"not_found":not_found,"uv":uv,"lag":lag,"elec":elec,"paper":paper}


def run_phase3(tracker_bytes, payment_bytes, log):
    log_line(log, "━━ PHASE 3 · STEPS 16–22 ━━━━━━━━━━━━━━━━", "ph")
    tracker_df = pd.read_excel(io.BytesIO(tracker_bytes), sheet_name="Timeline")
    total_rows = tracker_df.shape[0]

    try:
        pay = pd.read_csv(io.BytesIO(payment_bytes), low_memory=False)
        log_line(log, f"✓ Payment Report loaded · CSV · {len(pay):,} rows", "ok")
    except Exception:
        pay = pd.read_excel(io.BytesIO(payment_bytes))
        log_line(log, f"✓ Payment Report loaded · Excel · {len(pay):,} rows", "ok")

    rc = pay['PAYMENTDATE']
    if   pd.api.types.is_datetime64_any_dtype(rc): pay['PAYMENTDATE']=pd.to_datetime(rc,errors='coerce')
    elif pd.api.types.is_numeric_dtype(rc):         pay['PAYMENTDATE']=pd.to_datetime(rc,unit='D',origin='1899-12-30',errors='coerce')
    else:                                            pay['PAYMENTDATE']=pd.to_datetime(rc,errors='coerce')
    pay.loc[pay['PAYMENTDATE'].notna()&(pay['PAYMENTDATE']<MIN_VALID_DATE),'PAYMENTDATE']=pd.NaT
    pc = pay.dropna(subset=['STUDY_ID','PAYMENTDATE']).copy()
    pc['STUDY_ID'] = pc['STUDY_ID'].astype(str).str.strip()
    mp = pc.groupby('STUDY_ID')['PAYMENTDATE'].max().reset_index()
    plkp = dict(zip(mp['STUDY_ID'],mp['PAYMENTDATE']))
    log_line(log, f"✓ Payment pivot created · {len(plkp):,} unique Study IDs", "ok")

    wb = load_workbook(io.BytesIO(tracker_bytes)); ws = wb["Timeline"]
    ldr = total_rows + 1
    matched=updated=ay=ayu=ayk=ayn=nf=0

    for r in range(2, ldr+1):
        sr = ws.cell(row=r,column=1).value
        if sr is None: continue
        sid = str(sr).strip()
        if sid not in plkp: nf+=1; continue
        matched+=1; npd=plkp[sid]
        pr = ws.cell(row=r,column=28).value
        if str(pr).strip().upper()!="NO":
            ay+=1; ed=safe_to_datetime(ws.cell(row=r,column=29).value)
            if ed is not None:
                if npd.date()>ed.date(): ws.cell(row=r,column=29).value=to_python_datetime(npd); ayu+=1
                else: ws.cell(row=r,column=29).value=to_python_datetime(ed); ayk+=1
            else: ayn+=1
            continue
        er = ws.cell(row=r,column=26).value
        nr = "Payment" if (er is None or str(er).strip()=='') else f"Payment & {str(er).strip()}"
        ws.cell(row=r,column=26).value=nr; ws.cell(row=r,column=28).value="Yes"
        ed = safe_to_datetime(ws.cell(row=r,column=29).value)
        fd = (npd if ed is None or npd.date()>ed.date() else ed)
        ws.cell(row=r,column=29).value=to_python_datetime(fd); updated+=1

    for r in range(2, ldr+1):
        rc2=ws.cell(row=r,column=26)
        if rc2.value and str(rc2.value).strip(): rc2.alignment=Alignment(horizontal='center',vertical='center')
        dc=ws.cell(row=r,column=29)
        if dc.value and str(dc.value).strip(): dc.number_format='mm-dd-yy'; dc.alignment=Alignment(horizontal='center',vertical='center')

    log_line(log, f"✓ Matched: {matched:,} · NO→YES: {updated:,} · Already YES: {ay:,}", "ok")
    log_line(log, f"✓ Not in payment report: {nf:,}", "info")
    log_line(log, "✓ PHASE 3 COMPLETE → output ready for download", "ok")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read(), {"matched":matched,"updated":updated,"already_yes":ay,"not_found":nf}


def build_doclog_lookup(doclog_bytes, log):
    try:
        df = pd.read_csv(io.BytesIO(doclog_bytes), low_memory=False)
        log_line(log, f"✓ Doc-Log loaded · CSV · {len(df):,} rows", "ok")
    except Exception:
        df = pd.read_excel(io.BytesIO(doclog_bytes))
        log_line(log, f"✓ Doc-Log loaded · Excel · {len(df):,} rows", "ok")
    cols=list(df.columns); csid=cols[6]; csub=cols[8]; cdt=cols[19]
    df[csid]=df[csid].astype(str).str.strip()
    df[cdt]=pd.to_datetime(df[cdt],format='mixed',dayfirst=False,errors='coerce')
    valid=df.dropna(subset=[cdt]); mi=valid.groupby(csid)[cdt].idxmax(); dd=df.loc[mi]
    idx=dd.set_index(csid)
    afl=idx[csub].to_dict(); adt=idx[cdt].to_dict()
    adl={k: v.date() if pd.notna(v) else None for k,v in adt.items()}
    log_line(log, f"✓ Doc-Log lookup · {len(afl):,} unique Study IDs", "ok")
    return afl, adt, adl


def run_phase4(tracker_bytes, doclog_bytes, log):
    log_line(log, "━━ PHASE 4 · STEPS 24–34 ━━━━━━━━━━━━━━━━", "ph")
    tracker_df = pd.read_excel(io.BytesIO(tracker_bytes), sheet_name="Timeline")
    total_rows = tracker_df.shape[0]; ldr = total_rows+1

    afl, adt, adl = build_doclog_lookup(doclog_bytes, log)
    wb = load_workbook(io.BytesIO(tracker_bytes)); ws = wb["Timeline"]
    headers = [ws.cell(row=1,column=c).value for c in range(1,ws.max_column+1)]
    CA=get_col_idx(headers,'Study Id'); CV=get_col_idx(headers,'Bill Submission Date')
    CW=get_col_idx(headers,'Lag Time From Submission Date'); CZ=get_col_idx(headers,'Response Type')
    CAB=get_col_idx(headers,'Payment Received?'); CAC=get_col_idx(headers,'Last Payment Date')
    CAD=get_col_idx(headers,'EOR/Objection Received?'); CAE=get_col_idx(headers,'Last EOR/Objection Date')
    CTM=get_col_idx(headers,'Timely Response (Response in 60 Days)'); CAM=get_col_idx(headers,'SBR Sent Status')

    s26=s27=s28=ndl=s29=s29n=s29u=s30u=s30d=s30n=s30p=s30s=0
    s31v=s31f=s31p=s31n=s32=s32s=s34=s34s=0
    flagged=[]; sc=first_of_current_month()

    log_line(log, "→ Pass 1: Doc-Log lookups + SBR status...", "info")
    for r in range(2, ldr+1):
        sr=ws.cell(row=r,column=CA).value
        if sr is None: continue
        sid=str(sr).strip()
        if sid in afl:
            af=afl[sid]; ats=adt[sid]; afs=str(af).strip() if af is not None else ''; ado=to_date_obj(ats) if pd.notna(ats) else None
            if afs:
                zv=ws.cell(row=r,column=CZ).value; av=ws.cell(row=r,column=CAD).value; zs=str(zv).strip() if zv is not None else ''
                if normalize(av)=='no' and zs:
                    c=ws.cell(row=r,column=CZ); c.value=f"{zs} & {afs}"; c.alignment=Alignment(horizontal='center'); s26+=1
                elif not zs:
                    c=ws.cell(row=r,column=CZ); c.value=afs; c.alignment=Alignment(horizontal='center'); s27+=1
                av=ws.cell(row=r,column=CAD).value
                if normalize(av)=='no':
                    ws.cell(row=r,column=CAD).value="Yes"
                    dc=ws.cell(row=r,column=CAE); dc.value=ado; dc.number_format='M/D/YYYY'; dc.alignment=Alignment(horizontal='center'); s28+=1
            else: ndl+=1
        else: ndl+=1
        am=ws.cell(row=r,column=CAM).value; amn=normalize(am)
        if amn=='under billing cycle':
            ab=ws.cell(row=r,column=CAB).value; ad=ws.cell(row=r,column=CAD).value
            ac=ws.cell(row=r,column=CAC).value; ae=ws.cell(row=r,column=CAE).value
            rr=(normalize(ab)=='yes' or normalize(ad)=='yes')
            if rr:
                vd=[d for d in [to_date(ac),to_date(ae)] if d is not None]
                sol=max(vd)+timedelta(days=60) if vd else None
            else: sol=None
            if sol: ws.cell(row=r,column=CAM).value="PENDING SBR"; s29+=1
            else: s29n+=1
        else: s29u+=1
        am=ws.cell(row=r,column=CAM).value; amn=normalize(am)
        if amn in PROTECTED_STATUSES: s30p+=1; continue
        if amn not in PROCESS_STATUSES: s30s+=1; continue
        if sid not in afl: s30n+=1; continue
        af30=afl[sid]; ag30=adl[sid]
        if ag30 is None: s30n+=1; continue
        ae30=to_date(ws.cell(row=r,column=CAE).value)
        if ae30 is not None and ag30<=ae30: s30d+=1; continue
        write_date_cell(ws.cell(row=r,column=CAE),ag30)
        ws.cell(row=r,column=CAM).value="PENDING SBR"
        zv30=ws.cell(row=r,column=CZ).value
        nz=(str(zv30).strip()+" & "+str(af30).strip() if normalize(zv30) and contains_eob(zv30) else str(af30).strip())
        ws.cell(row=r,column=CZ).value=nz; s30u+=1

    log_line(log, f"✓ Steps 26-28: Z updated={s26+s27}, AD set={s28}", "ok")
    log_line(log, f"✓ Step 29: PENDING SBR set={s29}", "ok")
    log_line(log, f"✓ Step 30: Updated={s30u}", "ok")
    log_line(log, "→ Pass 2: SOL validation + Timely Response...", "info")

    for r in range(2, ldr+1):
        ab=ws.cell(row=r,column=CAB).value; ad=ws.cell(row=r,column=CAD).value
        ac=ws.cell(row=r,column=CAC).value; ae=ws.cell(row=r,column=CAE).value
        v=ws.cell(row=r,column=CV).value
        rr=(normalize(ab)=='yes' or normalize(ad)=='yes')
        acd=to_date(ac); aed=to_date(ae); vd=to_date(v)
        if rr:
            vds=[d for d in [acd,aed] if d is not None]
            sol=max(vds)+timedelta(days=60) if vds else None
        else: sol=None
        if sol is None: s31n+=1
        elif sol<sc: s31p+=1
        else:
            if normalize(ws.cell(row=r,column=CAM).value) in VALID_SOL_STATUSES: s31v+=1
            else:
                flagged.append((r,[ws.cell(row=r,column=c).value for c in range(1,ws.max_column+1)],sol,ws.cell(row=r,column=CAM).value))
                s31f+=1
        vr=[d for d in [acd,aed] if d is not None]
        if vd is not None and vr:
            diff=(max(vr)-vd).days
            if 1<=diff<=60: ws.cell(row=r,column=CTM).value="Yes"; ws.cell(row=r,column=CTM).alignment=Alignment(horizontal='center'); s32+=1
            else: s32s+=1
        else: s32s+=1
        tv=ws.cell(row=r,column=CTM).value
        if normalize(tv)!='no': s34s+=1; continue
        if not rr: s34s+=1; continue
        wv=ws.cell(row=r,column=CW).value; wn=to_number(wv)
        if wn is None and vd is not None: wn=(date.today()-vd).days
        wb2=(wv is None or (isinstance(wv,str) and wv.strip()==''))
        if not (wb2 or (wn is not None and wn<60)): s34s+=1; continue
        ws.cell(row=r,column=CTM).value="Yes"; ws.cell(row=r,column=CTM).alignment=Alignment(horizontal='center'); s34+=1

    if "Anomalies" in wb.sheetnames: del wb["Anomalies"]
    wa=wb.create_sheet("Anomalies")
    hf=Font(bold=True,color="FFFFFF"); hfl=PatternFill("solid",fgColor="C00000")
    ca=Alignment(horizontal='center',vertical='center'); ff=PatternFill("solid",fgColor="FFE0E0")
    for ci,h in enumerate(headers,1):
        c=wa.cell(row=1,column=ci,value=h); c.font=hf; c.fill=hfl; c.alignment=ca
    for or2,(ri,rd,sd,av) in enumerate(flagged,2):
        for ci,val in enumerate(rd,1):
            c=wa.cell(row=or2,column=ci,value=val); c.fill=ff
            if isinstance(val,datetime): c.number_format='M/D/YYYY'; c.alignment=Alignment(horizontal='center')

    log_line(log, f"✓ Step 31: Valid SOL={s31v} · Anomalies flagged={s31f}", "ok")
    log_line(log, f"✓ Step 32: Timely Response set={s32}", "ok")
    log_line(log, f"✓ Step 34: Anomaly fix={s34}", "ok")
    log_line(log, "✓ PHASE 4 COMPLETE → All phases done! ✓", "ok")

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read(), {"z":s26+s27,"ad":s28,"pending":s29,"step30":s30u,"anomalies":s31f,"timely":s32,"fix34":s34}


# ─────────────────────────────────────────────────────────────────────────────
# PIPELINE RUNNER
# ─────────────────────────────────────────────────────────────────────────────
def run_all_phases(tb, bb, sb, pb, db):
    log = st.session_state.run_log
    try:
        o1,s1 = run_phase1(tb,bb,log)
        st.session_state.p1_out=o1; st.session_state.p1_stats=s1
    except Exception as e:
        log_line(log,f"✗ PHASE 1 FAILED: {e}","error"); st.session_state.error_phase=1; return
    try:
        o2,s2 = run_phase2(o1,sb,log)
        st.session_state.p2_out=o2; st.session_state.p2_stats=s2
    except Exception as e:
        log_line(log,f"✗ PHASE 2 FAILED: {e}","error"); st.session_state.error_phase=2; return
    try:
        o3,s3 = run_phase3(o2,pb,log)
        st.session_state.p3_out=o3; st.session_state.p3_stats=s3
    except Exception as e:
        log_line(log,f"✗ PHASE 3 FAILED: {e}","error"); st.session_state.error_phase=3; return
    try:
        o4,s4 = run_phase4(o3,db,log)
        st.session_state.p4_out=o4; st.session_state.p4_stats=s4
    except Exception as e:
        log_line(log,f"✗ PHASE 4 FAILED: {e}","error"); st.session_state.error_phase=4; return
    log_line(log,"━━ ALL PHASES COMPLETE ✓ ━━━━━━━━━━━━━━━━━","ph")
    st.session_state.error_phase = None


# ─────────────────────────────────────────────────────────────────────────────
# UI HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def phase_status(n):
    ep = st.session_state.error_phase
    if ep and n==ep: return 'error'
    if st.session_state.get(f"p{n}_out"): return 'done'
    for i in range(1,5):
        if not st.session_state.get(f"p{i}_out"):
            return 'active' if i==n else 'idle'
    return 'idle'

def pip_status():
    p = [st.session_state.get(f"p{i}_out") is not None for i in range(1,5)]
    s = ['done','done' if p[0] else 'active',
         'done' if p[1] else ('active' if p[0] else 'idle'),
         'done' if p[2] else ('active' if p[1] else 'idle'),
         'done' if p[3] else ('active' if p[2] else 'idle')]
    return s

def fs_row(name, f):
    cls = 'ok' if f else 'miss'
    nc  = 'ok' if f else ''
    sz  = fmt_bytes(f.size) if f else '—'
    return f'<div class="fs-row"><div class="fs-dot {cls}"></div><span class="fs-name {nc}">{name}</span><span class="fs-sz">{sz}</span></div>'

def uz_card(f, name, ftype, wide=False):
    w = " wide" if wide else ""
    if f:
        return f"""<div class="uz-card ok{w}">
          <div class="uz-type {ftype}">{ftype.upper()}</div>
          <div class="uz-info">
            <div class="uz-name">{f.name}</div>
            <div class="uz-meta">✓ {fmt_bytes(f.size)} · Ready</div>
          </div><div class="uz-chk">✓</div></div>"""
    return f"""<div class="uz-card{w}">
      <div class="uz-type emp">+</div>
      <div class="uz-info">
        <div class="uz-name">Upload {name}</div>
        <div class="uz-meta">Use sidebar · Max 500 MB</div>
      </div></div>"""

def mv_html(v, cls=""):
    val = f"{v:,}" if isinstance(v,int) else str(v)
    return f'<div class="mc-val {cls}">{val}</div>'


# ─────────────────────────────────────────────────────────────────────────────
# NAVBAR
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="sbr-navbar">
  <div class="sbr-brand">
    <div class="sbr-brand-icon">SBR</div>
    <span class="sbr-brand-name">SBR Tracker Automation</span>
    <span class="sbr-brand-chip">v2.0</span>
  </div>
  <div class="sbr-nav-right">
    <div class="sbr-pill"><div class="sbr-dot green"></div>App ready</div>
    <div class="sbr-pill" style="color:#2e3448;">·</div>
    <div class="sbr-pill">Max file size: 500 MB</div>
  </div>
</div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<div class="sb-lbl">Upload Input Files</div>', unsafe_allow_html=True)

    tracker_file    = st.file_uploader("📂 Master-Tracker (.xlsx)",           type=["xlsx"],       key="f_t")
    bdr_file        = st.file_uploader("📂 BDR (.csv)",                       type=["csv"],        key="f_b")
    submission_file = st.file_uploader("📂 Submission Report (.xlsx / .csv)", type=["xlsx","csv"], key="f_s")
    payment_file    = st.file_uploader("📂 Payment Report (.csv / .xlsx)",    type=["csv","xlsx"], key="f_p")
    doclog_file     = st.file_uploader("📂 Doc-Log (.csv / .xlsx)",           type=["csv","xlsx"], key="f_d")

    # ── Python-side file size guard (500 MB) ─────────────────────────────────
    _oversized = []
    for _uf, _label in [
        (tracker_file,    "Master-Tracker"),
        (bdr_file,        "BDR"),
        (submission_file, "Submission Report"),
        (payment_file,    "Payment Report"),
        (doclog_file,     "Doc-Log"),
    ]:
        if _uf and _uf.size > MAX_FILE_BYTES:
            _oversized.append(f"{_label} ({fmt_bytes(_uf.size)})")
    if _oversized:
        st.error(f"⚠ File(s) exceed 500 MB limit: {', '.join(_oversized)}")
        if tracker_file    and tracker_file.size    > MAX_FILE_BYTES: tracker_file    = None
        if bdr_file        and bdr_file.size        > MAX_FILE_BYTES: bdr_file        = None
        if submission_file and submission_file.size > MAX_FILE_BYTES: submission_file = None
        if payment_file    and payment_file.size    > MAX_FILE_BYTES: payment_file    = None
        if doclog_file     and doclog_file.size     > MAX_FILE_BYTES: doclog_file     = None

    st.markdown('<div class="sb-lbl" style="margin-top:18px;">File Status</div>', unsafe_allow_html=True)
    st.markdown(
        fs_row("Master-Tracker.xlsx", tracker_file) +
        fs_row("BDR.csv",             bdr_file) +
        fs_row("Submission Report",   submission_file) +
        fs_row("Payment Report",      payment_file) +
        fs_row("Doc-Log",             doclog_file),
        unsafe_allow_html=True
    )

    st.markdown('<div class="sb-lbl" style="margin-top:18px;">Phase Status</div>', unsafe_allow_html=True)
    for n, title, sub in [(1,"Phase 1","Steps 1–6 · New Cases"),(2,"Phase 2","Steps 7–15 · Submission"),
                           (3,"Phase 3","Steps 16–22 · Payment"),(4,"Phase 4","Steps 24–34 · Doc-Log")]:
        ps = phase_status(n)
        icon = "✓" if ps=='done' else ("✗" if ps=='error' else str(n))
        badge = "done" if ps=='done' else ("error" if ps=='error' else ("running" if ps=='active' else "idle"))
        badge_txt = badge
        st.markdown(f"""
        <div class="phase-row {ps}">
          <div class="pnum {ps}">{icon}</div>
          <div class="pinfo"><div class="ptitle">{title}</div><div class="psub">{sub}</div></div>
          <span class="pbadge {ps}">{badge_txt}</span>
        </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
    if st.button("↺  Reset Everything", use_container_width=True, key="reset"):
        for k in ["p1_out","p2_out","p3_out","p4_out",
                  "p1_stats","p2_stats","p3_stats","p4_stats",
                  "run_log","error_phase"]:
            st.session_state[k] = [] if k=="run_log" else None
        st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# MAIN CONTENT
# ─────────────────────────────────────────────────────────────────────────────

# Header
st.markdown("""
<div class="main-hdr">
  <div class="main-title">Run Automation Pipeline</div>
  <div class="main-sub">Upload all 5 files · Click Run Once · All 4 phases chain automatically · Download each phase output as it completes</div>
</div>""", unsafe_allow_html=True)

# Pipeline progress bar
pip = pip_status()
labels = ["Upload","Phase 1","Phase 2","Phase 3","Phase 4"]
track  = "".join(
    f'<div class="pip-dot {pip[i]}"></div>' +
    (f'<div class="pip-line {"done" if i+1<5 and pip[i+1]=="done" else "idle"}"></div>' if i<4 else '')
    for i in range(5)
)
lbls = "".join(f'<span class="pip-lbl {pip[i]}">{labels[i]}</span>' for i in range(5))
st.markdown(f"""
<div class="pip-wrap">
  <div class="pip-track">{track}</div>
  <div class="pip-labels">{lbls}</div>
</div>""", unsafe_allow_html=True)

# Metric cards
s1 = st.session_state.p1_stats or {}
s3 = st.session_state.p3_stats or {}
s4 = st.session_state.p4_stats or {}
tr = s1.get("total_rows","—"); nc = s1.get("new_cases","—")
pf = s3.get("updated","—");    an = s4.get("anomalies","—")

st.markdown(f"""
<div class="metrics-row">
  <div class="mc">
    <div class="mc-lbl">Total Tracker Rows</div>
    {mv_html(tr)}
    <div class="mc-delta info">{f"+{s1['new_cases']:,} added this run" if s1 else "Run pipeline to see stats"}</div>
  </div>
  <div class="mc">
    <div class="mc-lbl">New Cases Added</div>
    {mv_html(nc,"blue")}
    <div class="mc-delta info">{"Phase 1 · Steps 1–6" if s1 else "Pending Phase 1"}</div>
  </div>
  <div class="mc">
    <div class="mc-lbl">Payments Flagged</div>
    {mv_html(pf,"green")}
    <div class="mc-delta">{"Phase 3 · Steps 16–22" if s3 else "Pending Phase 3"}</div>
  </div>
  <div class="mc">
    <div class="mc-lbl">Anomalies Found</div>
    {mv_html(an,"amber")}
    <div class="mc-delta">{"Phase 4 · Step 31" if s4 else "Pending Phase 4"}</div>
  </div>
</div>""", unsafe_allow_html=True)

# File status cards
all_files  = [tracker_file, bdr_file, submission_file, payment_file, doclog_file]
file_names = ["Master-Tracker.xlsx","BDR.csv","Submission Report","Payment-Report","Doc-Log"]
file_types = ["xl","csv","xl","csv","csv"]

r1 = uz_card(tracker_file,"Master-Tracker.xlsx","xl") + uz_card(bdr_file,"BDR.csv","csv")
r2 = uz_card(submission_file,"Submission Report","xl") + uz_card(payment_file,"Payment-Report","csv")
r3 = uz_card(doclog_file,"Doc-Log","csv",wide=True)

st.markdown(f"""
<div class="uz-outer">
  <div class="uz-title">Input Files &nbsp;·&nbsp;
    <span style="color:#2e3448;font-weight:400;font-size:9px;letter-spacing:0;text-transform:none;">
      Upload via sidebar · Max 500 MB per file
    </span>
  </div>
  <div class="uz-grid">{r1}</div>
  <div class="uz-grid" style="margin-top:8px;">{r2}</div>
  <div class="uz-grid" style="margin-top:8px;">{r3}</div>
</div>""", unsafe_allow_html=True)

# Run button
all_ready    = all(all_files)
missing_list = [file_names[i] for i,f in enumerate(all_files) if not f]

st.markdown('<div class="run-outer">', unsafe_allow_html=True)
run_clicked = st.button(
    "▶  Run All Phases  —  Steps 1–34 · Phases 1 → 2 → 3 → 4",
    key="run_all",
    disabled=not all_ready,
    use_container_width=True,
)
st.markdown('</div>', unsafe_allow_html=True)

if not all_ready:
    st.markdown(
        f'<div style="padding:2px 32px 10px;font-size:11px;color:#f59e0b;">⚠ Missing files: {", ".join(missing_list)}</div>',
        unsafe_allow_html=True
    )

if run_clicked and all_ready:
    for k in ["p1_out","p2_out","p3_out","p4_out","p1_stats","p2_stats","p3_stats","p4_stats","error_phase"]:
        st.session_state[k] = None
    st.session_state.run_log = []
    with st.spinner("Running all phases — this may take a few minutes for large files..."):
        run_all_phases(
            tracker_file.read(), bdr_file.read(),
            submission_file.read(), payment_file.read(), doclog_file.read()
        )
    st.rerun()

# Live log
run_log  = st.session_state.run_log or []
is_done  = st.session_state.p4_out is not None
has_err  = st.session_state.error_phase is not None

if run_log:
    if is_done:
        badge = '<span style="font-size:11px;color:#4ade80;margin-left:auto;">✓ All phases complete</span>'
    elif has_err:
        badge = f'<span style="font-size:11px;color:#f87171;margin-left:auto;">✗ Error in Phase {st.session_state.error_phase}</span>'
    else:
        pn = next((i for i in range(1,5) if not st.session_state.get(f"p{i}_out")),None)
        badge = (f'<div style="display:flex;align-items:center;gap:5px;margin-left:auto;">'
                 f'<div class="pulse"></div><span class="prun-txt">Phase {pn} running...</span></div>') if pn else ''

    body = "".join(run_log)
    st.markdown(f"""
    <div class="log-panel">
      <div class="log-hdr">
        <div class="log-hdr-dot"></div>
        <span class="log-hdr-title">Live Processing Log</span>
        {badge}
      </div>
      <div class="log-body">{body}</div>
    </div>""", unsafe_allow_html=True)

# Download cards
stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
mime  = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
dl_meta = [
    (1,"Phase 1","Steps 1–6",  "p1_out",f"Tracker_Phase1_{stamp}.xlsx"),
    (2,"Phase 2","Steps 7–15", "p2_out",f"Tracker_Phase2_{stamp}.xlsx"),
    (3,"Phase 3","Steps 16–22","p3_out",f"Tracker_Phase3_{stamp}.xlsx"),
    (4,"Phase 4","Steps 24–34","p4_out",f"Tracker_Phase4_{stamp}.xlsx"),
]

st.markdown('<div class="dl-outer"><div class="dl-title">Download Outputs — Each Phase Separately</div>', unsafe_allow_html=True)
cols = st.columns(5)

for i,(n,label,steps,key,fname) in enumerate(dl_meta):
    data = st.session_state.get(key)
    ep   = st.session_state.error_phase
    with cols[i]:
        cls  = "ready" if data else ("error" if ep==n else "")
        icon = "📊" if data else ("⚠️" if ep==n else "📊")
        nm   = steps if data else ("Failed" if ep==n else steps)
        sz   = fmt_bytes(len(data)) if data else "—"
        wait = '' if data else ('<div class="dl-wait">Waiting...</div>' if ep!=n else '<div class="dl-wait" style="color:#6b2828;">Error</div>')
        st.markdown(f"""
        <div class="dl-card {cls}">
          <div class="dl-lbl">{label}</div>
          <div class="dl-ico">{icon}</div>
          <div class="dl-name">{nm}</div>
          <div class="dl-sz">{sz}</div>
          {wait}
        </div>""", unsafe_allow_html=True)
        if data:
            st.download_button("⬇ Download", data=data, file_name=fname, mime=mime, key=f"dl_{n}", use_container_width=True)

# Final card
with cols[4]:
    fd = st.session_state.p4_out
    st.markdown(f"""
    <div class="dl-card {'final' if fd else ''}">
      <div class="dl-lbl">Final Output</div>
      <div class="dl-ico">🏁</div>
      <div class="dl-name">All Phases</div>
      <div class="dl-sz">{fmt_bytes(len(fd)) if fd else '—'}</div>
      {'<div class="dl-wait">After Phase 4</div>' if not fd else ''}
    </div>""", unsafe_allow_html=True)
    if fd:
        st.download_button("⬇ Download Final", data=fd, file_name=f"Tracker_FINAL_{stamp}.xlsx",
                           mime=mime, key="dl_final", use_container_width=True)

st.markdown('</div>', unsafe_allow_html=True)

# Footer
st.markdown("""
<div style="padding:14px 32px;border-top:1px solid #1c2030;margin-top:8px;">
  <span style="font-size:10px;color:#2e3448;font-family:'DM Mono',monospace;">
    SBR Tracker Automation · Steps 1–34 · Phases 1–4 · v2.0
  </span>
</div>""", unsafe_allow_html=True)
